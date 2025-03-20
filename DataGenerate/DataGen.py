import openpyxl

def dataGenerator():
    # li = [['uname1','pass1'],['uname2','pass2'],['uname3','pass3']]

    li =[]
    li1 =[]
    wb = openpyxl.load_workbook(r"C:\Users\Niru\Desktop\Codes\New folder\UserList.xlsx")
    sh = wb['Sheet1']
    r = sh.max_row
    for i in range(1,r+1):
        li1=[]
        un = sh.cell(i,1)
        up = sh.cell(i,2)
        li1.insert(0, un.value)
        li1.insert(1, up.value)
        li.insert(i-1,li1)

    print(li)
    return li